from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os
import sys
import re


filename_raport="./asz.xlsx"


if os.path.exists(filename_raport):
    workbook = load_workbook(filename=filename_raport)
    sheet_ac = workbook.get_sheet_by_name('azt')
    workbook.active=sheet_ac
    sheet = workbook.active
else:
    workbook = Workbook()
    workbook.create_sheet('azt')
    sheet_ac = workbook.get_sheet_by_name('azt')
    workbook.active=sheet_ac
    sheet = workbook.active
    sheet["A1"] = "Recordings"
    sheet.column_dimensions['A'].width =13



sheet.insert_rows(idx=2)

sheet["A2"] = "adsfsfd"

workbook.save(filename=filename_raport)
