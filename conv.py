from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os
import sys
import re
from openpyxl.styles import PatternFill, colors, Alignment
import ast
import pickle

import copy

dict_legend_path="./dict_legend.txt"

infile = r"./test/path/jmeter.log"

JSON_RAPORT_PATH=sys.argv[2]

JSON_CONFIG_PATH=sys.argv[4]+"/.config"
JSON_RDB_INFO=sys.argv[5]

active_line="6"
col_nr_Recordings='A'
col_nr_How_many_rec_should_percent='B'


col_nr_Trace_suc='C'
col_nr_Trace_err='D'

col_nr_Trace_percent='E'
col_nr_TraceSpan='F'
col_nr_JMCount='G'
col_nr_Calls='H'
col_nr_Sent='I'
col_nr_Response='J'
col_nr_Successes='K'
col_nr_Module='L'
col_nr_Delay='M'


def hidden_col(worksheet):
    worksheet.column_dimensions[col_nr_Trace_suc].hidden= True
    worksheet.column_dimensions[col_nr_Trace_err].hidden= True
    worksheet.column_dimensions[col_nr_Sent].hidden= True
    worksheet.column_dimensions[col_nr_Response].hidden= True



def style(sheet,active_line):
    color = str(getColor(getMode(str(sys.argv[7])) ))
    Line_legend="6"
    setStyl(sheet[col_nr_Recordings+active_line],color)

    setStyl(sheet[col_nr_Trace_suc+active_line],color)
    setStyl(sheet[col_nr_Trace_err+active_line],color)
    setStyl(sheet[col_nr_Trace_percent+active_line],color)
    
    setStyl(sheet[col_nr_TraceSpan+active_line],color)
    setStyl(sheet[col_nr_JMCount+active_line],color)
    setStyl(sheet[col_nr_Calls+active_line],color)
    setStyl(sheet[col_nr_Sent+active_line],color)
    setStyl(sheet[col_nr_Response+active_line],color)
    setStyl(sheet[col_nr_Successes+active_line],color)
    setStyl(sheet[col_nr_Module+active_line],color)
    setStyl(sheet[col_nr_Delay+active_line],color)
    setStyl(sheet[col_nr_How_many_rec_should_percent+active_line],color)
    


def setStyl(call,color):
    call.fill = PatternFill(fgColor=color, fill_type = "solid")
    
def getMode(x):
    return {
        '1': "rdb and apm",
        '2': "apm",
        '3': "none",
        '4': "rdb"
    }[x]

def getColor(x):
    return {
        'rdb and apm': "ffdbd9",
        'apm': "00CCFFFF",
        'none': "00FFFFCC",
        'rdb': "00CCFFCC"
    }[x]

def createSheet(sheet):
    Line_RevDeBug_spec="1"
    Line_Application_spec="2"
    Line_Application_name="3"
    Line_legend="5"

    sheet["A"+Line_RevDeBug_spec] = "RevDeBug spec"
    sheet["B"+Line_RevDeBug_spec] = "CPU: "
    sheet["C"+Line_RevDeBug_spec] = rdb_info['cpu']
    sheet["D"+Line_RevDeBug_spec] = "RAM: " 
    sheet["E"+Line_RevDeBug_spec] =  rdb_info['ram']
    sheet["F"+Line_RevDeBug_spec] = "Size: "
    sheet["G"+Line_RevDeBug_spec] =  rdb_info['size']

    sheet["A"+Line_Application_spec] = "Application spec"
    sheet["B"+Line_Application_spec] = "CPU: "
    sheet["C"+Line_Application_spec] = json_config['cpu']
    sheet["D"+Line_Application_spec] = "RAM: " 
    sheet["E"+Line_Application_spec] =  json_config['ram']
    sheet["F"+Line_Application_spec] = "Size: "
    sheet["G"+Line_Application_spec] =  json_config['size']

    sheet["A"+Line_Application_name] = "Application name"
    sheet["B"+Line_Application_name] = json_config['application_name']
    sheet["C"+Line_Application_name] = "Git link"
    sheet["D"+Line_Application_name] = json_config['git_link'] 
    sheet["E"+Line_Application_name] = "RevDeBug server"
    sheet["F"+Line_Application_name] = json_config['server_rdb']
    sheet["G"+Line_Application_name] = "Language"
    sheet["H"+Line_Application_name] = json_config['language']
 

    sheet[col_nr_Trace_suc+Line_legend] = "Trace successes"
    sheet[col_nr_Trace_err+Line_legend] = "Trace error"
    sheet[col_nr_Trace_percent+Line_legend] = "Perc trace"


    sheet[col_nr_Recordings+Line_legend] = "Recordings"
    sheet[col_nr_TraceSpan+Line_legend] = "Trace Span"
    sheet[col_nr_JMCount+Line_legend] = "JM Count"
    sheet[col_nr_Calls+Line_legend] = "Calls/s avg" 
    sheet[col_nr_Sent+Line_legend] = "Sent  kb/s"
    sheet[col_nr_Response+Line_legend] = "Response AVG /s"
    sheet[col_nr_Successes+Line_legend] = "Successes set/real"
    sheet[col_nr_Module+Line_legend] = "RDB/APM"
    sheet[col_nr_Delay+Line_legend] = "Delay"
    # sheet[col_nr_JmeterErr+Line_legend] = "Jmeter proportion"
    sheet[col_nr_How_many_rec_should_percent+Line_legend] = "Perc recordings"

    sheet[col_nr_Recordings+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_TraceSpan+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_JMCount+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_Calls+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_Sent+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_Response+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_Successes+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_Module+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_Delay+Line_legend].alignment = Alignment(wrap_text=True)
    # sheet[col_nr_JmeterErr+Line_legend].alignment = Alignment(wrap_text=True)
    sheet[col_nr_How_many_rec_should_percent+Line_legend].alignment = Alignment(wrap_text=True)
    

    sheet.column_dimensions[col_nr_Recordings].width =10
    sheet.column_dimensions[col_nr_TraceSpan].width =10
    sheet.column_dimensions[col_nr_JMCount].width =10
    sheet.column_dimensions[col_nr_Calls].width =10
    sheet.column_dimensions[col_nr_Sent].width =10
    sheet.column_dimensions[col_nr_Response].width =15
    sheet.column_dimensions[col_nr_Successes].width =15
    sheet.column_dimensions[col_nr_Module].width =15
    sheet.column_dimensions[col_nr_Delay].width =15
    # sheet.column_dimensions[col_nr_JmeterErr].width =15
    sheet.column_dimensions[col_nr_How_many_rec_should_percent].width =12



with open(JSON_RAPORT_PATH) as f:
    json_dict = json.load(f)

with open(JSON_CONFIG_PATH) as f:
    json_config = json.load(f)

with open(JSON_RDB_INFO) as f:
    rdb_info = json.load(f)

describe=''
if "raport_name" in json_config:
    if not (json_config['raport_name'] is None):
        describe=describe+json_config['raport_name']+"_"

filename_raport="raports/"+json_config['language']+"_"+describe+sys.argv[3]

sheet_name=json_config['application_name'] 

with open(infile) as f:
    f = f.readlines()

calls=0
list_result=[]
for line in f: 
    x = re.search("summary = .* Avg", line)
    if x:
        tmp =x.group(0).split("=")
        tmp[2]=tmp[2][:-3]
        list_result.append(tmp[2])
        calls=list_result[len(list_result)-1].split("/s")[0]





if os.path.exists(filename_raport):
    workbook = load_workbook(filename=filename_raport)
    if not sheet_name in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
        sheet_ac = workbook[sheet_name]
        workbook.active=sheet_ac
        sheet = workbook.active
        createSheet(sheet)
    else:
        sheet_ac = workbook[sheet_name]
        workbook.active=sheet_ac
        sheet = workbook.active
    
else:
    workbook = Workbook()
    workbook.create_sheet(sheet_name)
    sheet_ac = workbook[sheet_name]
    workbook.active=sheet_ac
    sheet = workbook.active
    createSheet(sheet)

sheet.insert_rows(idx=6)


def percent_prop_rec(successes_percent, jm_count, value):
    How_many_should=((100-int(successes_percent))/100)*jm_count
    if(int(How_many_should)!=0):
        result =int((int(value)/How_many_should)*100)
        result=str(result)+ "%"
    else:
        result="-"
    return result

def percent_prop_trace(successes_percent, jm_count, value):
    if(int(jm_count)!=0):
        result =int((int(value)/jm_count)*100)
        result=str(result)+ "%"
    else:
        result="-"
    return result

successes=str(sys.argv[1])
data_name=successes+'/'+str(100-int(float(json_dict["TotalJmeter"]["errorPct"])))

value={}
value["Recordings"]=str(int(float(json_dict["Recordings"])))
value["TraceSpan"] = int(float(json_dict["Trace_Span"]))
value['JMCount']= int(float(json_dict["TotalJmeter"]["sampleCount"]))
value['Calls']= int(float(calls.strip()))
value['Sent']= int(float(json_dict["TotalJmeter"]["sentKBytesPerSec"]))
value['Response']= int(float(json_dict["TotalJmeter"]["meanResTime"]))
value['Successes']= data_name



value["Trace_err"]=int(float(json_dict["Trace_Span_Err"]))
value["Trace_suc"]=int(float(json_dict["Trace_Span_Suc"]))

# value['JmeterErr'] =int(float(json_dict["TotalJmeter"]["errorPct"]))
value['Module']= getMode(str(sys.argv[7]))
value['Delay']= sys.argv[6]

if (str(sys.argv[7])=="4" or str(sys.argv[7])=="1"):
    value['How_many_rec_should_percent']=percent_prop_rec(successes,value['JMCount'],value["Recordings"])
else:
    value['How_many_rec_should_percent']="-"

if (str(sys.argv[7])=="2" or str(sys.argv[7])=="1"):
    value["Trace_percent"]=percent_prop_trace(successes,value['JMCount'],value["TraceSpan"])
else:
    value["Trace_percent"]="-"


            
style(sheet, active_line)
sheet[col_nr_Recordings+active_line]  = value["Recordings"]

sheet[col_nr_Trace_err+active_line]  = value["Trace_err"]
sheet[col_nr_Trace_suc+active_line]  = value["Trace_suc"]
sheet[col_nr_Trace_percent+active_line]  = value["Trace_percent"]

sheet[col_nr_TraceSpan+active_line]  = value["TraceSpan"]
sheet[col_nr_JMCount+active_line] = value['JMCount']
sheet[col_nr_Calls+active_line]=value['Calls']
sheet[col_nr_Sent+active_line] = value['Sent']
sheet[col_nr_Response+active_line] = value['Response']
sheet[col_nr_Successes+active_line] =value['Successes']
sheet[col_nr_Module+active_line] = value['Module'] 
sheet[col_nr_Delay+active_line] = value['Delay']
# sheet[col_nr_JmeterErr+active_line]= value['JmeterErr']
sheet[col_nr_How_many_rec_should_percent+active_line]= value['How_many_rec_should_percent']

hidden_col(sheet)

workbook.save(filename=filename_raport)
