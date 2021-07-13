from openpyxl import Workbook
from openpyxl import load_workbook
import json
import os
import sys
JSON_RES="rdb_res.json"
filename_raport=sys.argv[1]

# data deletion

def createSheet(sheet,f):
    sheet["A1"] = "Removed:"
    sheet["B1"] = f[0]

    sheet["A2"] = "within (seconds):"
    sheet["B2"] = f[1]

    sheet.column_dimensions['A'].width =15
    sheet.column_dimensions['B'].width =15

f = open("rdb_res", "r")
f=f.read().splitlines()


sheet_name="RDB data"


if os.path.exists(filename_raport):
    workbook = load_workbook(filename=filename_raport)
    workbook.create_sheet(sheet_name)
    sheet_ac = workbook.get_sheet_by_name(sheet_name)
    workbook.active=sheet_ac
    sheet = workbook.active
    createSheet(sheet,f)   
else:
    workbook = Workbook()
    workbook.create_sheet(sheet_name)
    sheet_ac = workbook.get_sheet_by_name(sheet_name)
    workbook.active=sheet_ac
    sheet = workbook.active
    createSheet(sheet,f)

workbook.save(filename=filename_raport)
